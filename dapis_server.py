# dapis_server.py
from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, JSONResponse
from pydantic import BaseModel
import re
import os
from pathlib import Path
import json
import sqlite3
from concurrent.futures import ProcessPoolExecutor
import multiprocessing
import atexit
import openpyxl
import fitz  # PyMuPDF
import platform
import subprocess
import argparse

VERSION = "0.1.1"
PROGRAM = "Dapis Server"
LIBREOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"

app = FastAPI()
SESSIONS = {}

# -----------------------------
# 設定管理
# -----------------------------
def load_config_file(file_path: str) -> list[str]:
    argv = []
    if not os.path.exists(file_path):
        return argv
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            argv.extend(line.split())
    return argv

def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Dapis Server")
    parser.add_argument("--host", type=str, help="Server host")
    parser.add_argument("--port", type=int, help="Server port")
    parser.add_argument("--results_dir", type=str, help="Directory for SQLite DBs and results")
    parser.add_argument("--targets_count_default", type=int, help="Default number of targets to display in browser")
    parser.add_argument("--config_file", type=str, help="Path to config file", default=None)
    return parser.parse_args(argv)

class Config:
    DEFAULTS = {
        "host": "127.0.0.1",
        "port": 8000,
        "results_dir": "results",
        "targets_count_default": 99,
    }
    ENV_VARS = {
        "host": "DAPIS_HOST",
        "port": "DAPIS_PORT",
        "results_dir": "DAPIS_RESULTS_DIR",
        "targets_count_default": "DAPIS_TARGETS_COUNT_DEFAULT",
    }
    def __init__(self):
        self.args = parse_args()
        file_args = []
        if self.args.config_file:
            file_args = load_config_file(self.args.config_file)
        self.file_args = parse_args(file_args)

    def get(self, key: str):
        val = getattr(self.args, key, None)
        if val is not None:
            return val
        val = getattr(self.file_args, key, None)
        if val is not None:
            return val
        env_var = self.ENV_VARS.get(key)
        if env_var and env_var in os.environ:
            env_val = os.environ[env_var]
            if isinstance(self.DEFAULTS[key], int):
                try:
                    return int(env_val)
                except ValueError:
                    pass
            return env_val
        return self.DEFAULTS[key]

cfg = Config()
RESULTS_DIR = cfg.get("results_dir")
TARGETS_COUNT_DEFAULT = cfg.get("targets_count_default")

# -----------------------------
# プロセスプール・終了処理
# -----------------------------
executor = ProcessPoolExecutor(max_workers=multiprocessing.cpu_count())

def cleanup():
    global executor
    if executor:
        executor.shutdown(wait=True)
    for p in multiprocessing.active_children():
        p.terminate()
        p.join()

atexit.register(cleanup)

# -----------------------------
# モデル
# -----------------------------
class Targets(BaseModel):
    session_id: str
    paths: list[str]

# -----------------------------
# セッション管理
# -----------------------------
@app.post("/submit_targets")
async def submit_targets(data: Targets):
    SESSIONS[data.session_id] = data.paths
    return {"ok": True, "session_id": data.session_id}

@app.get("/get_targets")
async def get_targets(session_id: str):
    if session_id not in SESSIONS:
        return JSONResponse({"error":"unknown session_id"}, status_code=400)
    # 表示件数制御
    max_count = TARGETS_COUNT_DEFAULT
    paths = SESSIONS[session_id]
    displayed = paths[:max_count]
    if len(paths) > max_count:
        displayed.append(f"...and {len(paths)-max_count} more")
    return {"paths": displayed}

# -----------------------------
# SQLite DB
# -----------------------------
def init_db(session_id: str):
    Path(RESULTS_DIR).mkdir(exist_ok=True)
    db_path = Path(RESULTS_DIR) / f"{session_id}.sqlite"
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS results (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        query TEXT,
        file_path TEXT,
        sheet TEXT,
        line TEXT,
        column TEXT,
        page TEXT,
        value TEXT
    )
    """)
    conn.commit()
    return conn

def save_result(conn, query, file_path, sheet, line, column, page, value):
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO results (query, file_path, sheet, line, column, page, value)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (query, file_path, sheet, line, column, page, value))
    conn.commit()

# --- 検索関数 ---
def search_text_file(fp, pattern, flags):
    rx = re.compile(pattern, flags)
    out = []
    try:
        with open(fp, "r", encoding="utf-8", errors="ignore") as fh:
            for i, line in enumerate(fh, start=1):
                if rx.search(line):
                    out.append({
                        "path": fp,
                        "sheet": None,
                        "line": str(i),
                        "column": None,
                        "page": None,
                        "value": line.rstrip("\n")
                    })
    except Exception:
        pass
    return out

def search_excel_file(fp, pattern, flags):
    rx = re.compile(pattern, flags)
    out = []
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for j, cell in enumerate(row, start=1):
                    val = str(cell) if cell is not None else ""
                    if rx.search(val):
                        out.append({
                            "path": fp,
                            "sheet": sheet,
                            "line": str(i),
                            "column": str(j),
                            "page": None,
                            "value": val
                        })
    except Exception:
        pass
    return out

def convert_word_to_pdf(input_path: str, output_path: str = None) -> str:
    """
    WordファイルをPDFに変換する (OSごとに自動判別)
    - Windows: Word COM API
    - macOS/Linux: LibreOffice (headless mode)
    """
    if output_path is None:
        output_path = os.path.splitext(input_path)[0] + ".pdf"

    system = platform.system()

    if system == "Windows":
        import comtypes.client
        word = comtypes.client.CreateObject("Word.Application")
        doc = word.Documents.Open(input_path)
        doc.SaveAs(output_path, FileFormat=17)  # 17 = wdFormatPDF
        doc.Close()
        word.Quit()
    else:
        subprocess.run([
            LIBREOFFICE, "--headless", "--convert-to", "pdf", "--outdir",
            os.path.dirname(output_path), input_path
        ], check=True)

    return output_path

def convert_ppt_to_pdf(input_path: str, output_path: str = None) -> str:
    """
    PowerPointファイルをPDFに変換する
    Windowsでは COM API を優先、失敗時に LibreOffice を利用
    macOS/Linuxでは LibreOffice のみ
    """

    if output_path is None:
        output_path = os.path.splitext(input_path)[0] + ".pdf"

    system = platform.system()

    # --- Windows: COM API を優先 ---
    if system == "Windows":
        try:
            import comtypes.client
            powerpoint = comtypes.client.CreateObject("Powerpoint.Application")
            powerpoint.Visible = 1
            presentation = powerpoint.Presentations.Open(input_path)
            presentation.SaveAs(output_path, 32)  # 32 = PDF format
            presentation.Close()
            powerpoint.Quit()
            return output_path
        except Exception as e:
            print(f"[WARN] PowerPoint COM API failed: {e}. Trying LibreOffice...")

    # --- LibreOffice fallback (全OS対応) ---
    try:
        subprocess.run([
            LIBREOFFICE, "--headless", "--convert-to", "pdf", input_path,
            "--outdir", os.path.dirname(output_path)
        ], check=True)
        return output_path
    except Exception as e:
        raise RuntimeError(f"Failed to convert PowerPoint to PDF: {e}")

def search_pdf_file(fp, pattern, flags, fp_alias=None):
    rx = re.compile(pattern, flags)
    out = []
    try:
        if fp_alias is None:
            fp_alias = fp
        doc = fitz.open(fp)
        for page_num, page in enumerate(doc, start=1):
            words = page.get_text("words")  # 各単語ごとの位置情報
            for w in words:
                x0, y0, x1, y1, word, block_no, line_no, word_no = w
                if rx.search(word):
                    # ページサイズ取得
                    page_width, page_height = page.rect.width, page.rect.height
                    # x%, y% 計算
                    x_pct = (x0 / page_width) * 100
                    y_pct = (y0 / page_height) * 100
                    out.append({
                        "path": fp_alias,
                        "sheet": None,
                        "line": str(line_no),
                        "column": None,
                        "page": str(page_num),
                        "value": word,
                        "x%": round(x_pct, 1),
                        "y%": round(y_pct, 1)
                    })
    except Exception:
        pass
    return out

def search_file(fp, pattern, flags):
    ext = fp.lower().split(".")[-1]
    if ext in ("xlsx", "xlsm", "xls"):
        return search_excel_file(fp, pattern, flags)
    elif ext in ("docx"):
        word_fp = fp
        pdf_fp = convert_word_to_pdf(word_fp)
        return search_pdf_file(pdf_fp, pattern, flags, word_fp)
    elif ext in ("pptx"):
        ppt_fp = fp
        pdf_fp = convert_ppt_to_pdf(ppt_fp)
        return search_pdf_file(pdf_fp, pattern, flags, ppt_fp)
    elif ext == "pdf":
        return search_pdf_file(fp, pattern, flags)
    elif ext in ("txt", "py", "md", "csv"):
        return search_text_file(fp, pattern, flags)
    else:
        return []

@app.post("/search")
async def search(req: Request):
    body = await req.json()
    session_id = body.get("session_id")
    pattern = body.get("pattern")
    ignore_case = body.get("ignore_case", False)

    if session_id not in SESSIONS:
        return JSONResponse({"error": "unknown session_id"}, status_code=400)

    conn = init_db(session_id)
    flags = re.IGNORECASE if ignore_case else 0

    paths = SESSIONS[session_id]
    futures = [executor.submit(search_file, fp, pattern, flags) for fp in paths]
    results = []
    for f in futures:
        results.extend(f.result())

    # SQLite に保存
    for r in results:
        save_result(conn, pattern, r["path"], r["sheet"], r["line"], r["column"], r["page"], r["value"])

    return {"matches": results, "version": VERSION, "program": PROGRAM}

@app.get("/", response_class=HTMLResponse)
async def index():
    return f"""
<!doctype html>
<html>
<head><meta charset="utf-8"><title>Dapis Text Search Powered by {PROGRAM} version:{VERSION}</title></head>
<body>
<h2>Dapis Text Search Powered by {PROGRAM} version:{VERSION}</h2>
<form id="f" onsubmit="return false;">
  <input type="hidden" id="session_id" value="" />
  <div>Search Paths:</div>
  <pre id="file_list"></pre>
  <div>
    Pattern (regex): <input name="pattern" id="pattern"/><br/>
    Ignore case: <input type="checkbox" id="ic" /><br/>
    <button type="button" onclick="doSearch()">Search</button>
  </div>
</form>
<pre id="out"></pre>

<script>
async function fetchTargets() {{
  const sid = new URL(location.href).searchParams.get("session_id") || "";
  if (!sid) return;
  document.getElementById("session_id").value = sid;
  const res = await fetch(`/get_targets?session_id=${{sid}}`);
  if (!res.ok) return;
  const data = await res.json();
  document.getElementById("file_list").textContent = data.paths.join("\\n");
}}

async function doSearchInternal(){{
  const sid = document.getElementById('session_id').value;
  const pat = document.getElementById('pattern').value;
  const ic = document.getElementById('ic').checked;
  const res = await fetch('/search', {{
    method:'POST',
    headers:{{'Content-Type':'application/json'}},
    body: JSON.stringify({{session_id:sid, pattern:pat, ignore_case:ic}})
  }});
  const j = await res.json();
  document.getElementById('out').textContent = JSON.stringify(j, null, 2);
}}

let searching = false;
async function doSearch(){{
  if(searching) return;
  searching = true;

  const btn = document.querySelector('button[onclick="doSearch()"]');
  const out = document.getElementById('out');

  // ボタン無効化＆表示変更
  btn.disabled = true;
  btn.textContent = "検索中…";
  out.textContent = "";

  try {{
    await doSearchInternal();
  }} catch(e) {{
    out.textContent = "検索エラー: " + e;
  }} finally {{
    // ボタン有効化＆表示を元に戻す
    btn.disabled = false;
    btn.textContent = "Search";
    searching = false;
  }}
}}

window.onload = fetchTargets;
</script>
</body>
</html>
"""

if __name__ == "__main__":
    import uvicorn
    import pathlib
    _host = "127.0.0.1"
    _port = 8000
    module_name = pathlib.Path(__file__).stem.replace("-", "_")
    uvicorn.run(f"{module_name}:app", host=_host, port=_port, reload=False)

